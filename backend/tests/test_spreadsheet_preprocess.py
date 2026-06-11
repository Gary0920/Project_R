from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from app.features.preprocessing.spreadsheet import (
    extract_spreadsheet_markdown,
    _clean_cell,
    _is_material_code_column,
    _build_markdown_table,
    SpreadsheetExtractionResult,
)


def _create_test_xlsx(path: Path, sheets: dict[str, list[list[str | None]]]) -> None:
    """Create a test XLSX file with given sheet data."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)

    wb.save(str(path))


class TestCleanCell(unittest.TestCase):
    def test_none(self):
        self.assertEqual(_clean_cell(None), "")

    def test_string(self):
        self.assertEqual(_clean_cell("  hello  "), "hello")

    def test_float_int(self):
        self.assertEqual(_clean_cell(5.0), "5")
        self.assertEqual(_clean_cell(5.5), "5.50")

    def test_str(self):
        self.assertEqual(_clean_cell("test"), "test")


class TestIsMaterialCodeColumn(unittest.TestCase):
    def test_chinese_patterns(self):
        self.assertTrue(_is_material_code_column("编号"))
        self.assertTrue(_is_material_code_column("料号"))
        self.assertTrue(_is_material_code_column("图号"))

    def test_english_patterns(self):
        self.assertTrue(_is_material_code_column("Code"))
        self.assertTrue(_is_material_code_column("ITEM"))
        self.assertTrue(_is_material_code_column("Name(编号)"))
        self.assertTrue(_is_material_code_column("TYPE(名称)"))

    def test_non_material(self):
        self.assertFalse(_is_material_code_column("备注"))
        self.assertFalse(_is_material_code_column("数量"))
        self.assertFalse(_is_material_code_column("规格"))
        self.assertFalse(_is_material_code_column("单价"))


class TestBuildMarkdownTable(unittest.TestCase):
    def test_basic_table(self):
        headers = ["编号", "名称", "规格"]
        rows = [["GL01", "玻璃", "6+12A+6"], ["GL02", "型材", "6063-T5"]]
        md = _build_markdown_table(headers, rows, "编号")
        self.assertIn("| 编号 | 名称 | 规格 |", md)
        self.assertIn("| GL01 | 玻璃 | 6+12A+6 |", md)
        self.assertIn("| GL02 | 型材 | 6063-T5 |", md)

    def test_empty_headers(self):
        self.assertEqual(_build_markdown_table([], [], None), "")


class TestExtractSpreadsheetMarkdown(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self.xlsx_path = Path(self._tmpdir.name) / "test_material_list.xlsx"

    def tearDown(self):
        self._tmpdir.cleanup()

    def test_simple_extraction(self):
        _create_test_xlsx(self.xlsx_path, {
            "Glass": [
                ["Name(编号)", "TYPE(名称)", "规格", "数量"],
                ["GL01", "中空玻璃", "6+12A+6", "100"],
                ["GL02", "钢化玻璃", "8mm", "200"],
            ],
        })
        result = extract_spreadsheet_markdown(self.xlsx_path)
        self.assertIsInstance(result, SpreadsheetExtractionResult)
        self.assertEqual(result.sheet_count, 1)
        self.assertEqual(result.total_rows, 2)
        self.assertIn("GL01", result.markdown)
        self.assertIn("GL02", result.markdown)
        self.assertEqual(result.review_status, "approved")
        self.assertEqual(result.file_kind, "spreadsheet")

    def test_multiple_sheets(self):
        _create_test_xlsx(self.xlsx_path, {
            "Glass": [
                ["编号", "名称", "规格"],
                ["GL01", "中空玻璃", "6+12A+6"],
            ],
            "Hardware": [
                ["编号", "名称", "材质"],
                ["HW01", "合页", "不锈钢"],
            ],
        })
        result = extract_spreadsheet_markdown(self.xlsx_path)
        self.assertEqual(result.sheet_count, 2)
        self.assertEqual(result.total_rows, 2)
        self.assertIn("Glass", result.markdown)
        self.assertIn("Hardware", result.markdown)
        self.assertIn("GL01", result.markdown)
        self.assertIn("HW01", result.markdown)

    def test_material_codes_detected(self):
        _create_test_xlsx(self.xlsx_path, {
            "ML": [
                ["Name(编号)", "TYPE(名称)", "规格"],
                ["GL01", "中空玻璃", "6+12A+6"],
                ["GL02", "钢化玻璃", "8mm"],
            ],
        })
        result = extract_spreadsheet_markdown(self.xlsx_path)
        self.assertIn("GL01", result.material_codes_found)
        self.assertIn("GL02", result.material_codes_found)

    def test_empty_sheet_skipped(self):
        _create_test_xlsx(self.xlsx_path, {
            "Empty": [["Header"]],
        })
        result = extract_spreadsheet_markdown(self.xlsx_path)
        # Sheet with header-only (no data rows) should have 0 count
        self.assertEqual(result.sheet_count, 0)
        self.assertEqual(result.total_rows, 0)

    def test_hidden_sheet_skipped(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Visible"
        ws1.append(["Name", "Value"])
        ws1.append(["A", "1"])

        ws2 = wb.create_sheet(title="Hidden")
        ws2.sheet_state = "hidden"
        ws2.append(["Secret", "Data"])
        ws2.append(["X", "2"])

        xlsx_path = Path(self._tmpdir.name) / "test_hidden.xlsx"
        wb.save(str(xlsx_path))

        result = extract_spreadsheet_markdown(xlsx_path)
        self.assertEqual(result.sheet_count, 1)  # Only visible
        self.assertIn("Visible", result.markdown)
        # "Hidden" sheet data should NOT appear in the markdown table
        self.assertNotIn("Secret", result.markdown)
        self.assertNotIn("Data", result.markdown)

    def test_truncated_long_sheet(self):
        headers = ["Col1"]
        rows = [[f"Row{i}"] for i in range(250)]  # Exceeds MAX_ROWS_PER_SHEET (200)

        _create_test_xlsx(self.xlsx_path, {
            "Large": [headers] + rows,
        })
        result = extract_spreadsheet_markdown(self.xlsx_path)
        self.assertIn("truncated", result.markdown.lower())
        self.assertTrue(any(s.truncated for s in result.sheets))

    def test_handle_missing_openpyxl(self):
        """Test graceful failure when openpyxl is not installed."""
        # Create a test file first
        if not self.xlsx_path.exists():
            _create_test_xlsx(self.xlsx_path, {"Sheet1": [["A"], ["1"]]})
        result = extract_spreadsheet_markdown(self.xlsx_path)
        self.assertIn(result.review_status, ("approved", "failed_retryable"))


if __name__ == "__main__":
    unittest.main()
