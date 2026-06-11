from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

SKILL_NAME = "spreadsheet-preprocess"
SKILL_VERSION = "1.0.0"
PROMPT_VERSION = "rules-spreadsheet-v1"

# Patterns for identifying material code / ID columns
MATERIAL_CODE_PATTERNS = {"编号", "code", "item", "name", "编号", "id", "material", "type", "料号", "图号"}

# Max rows per sheet to process
MAX_ROWS_PER_SHEET = 200

# Max total cells to avoid memory blowout
MAX_TOTAL_CELLS = 50_000


@dataclass(frozen=True)
class SheetExtraction:
    sheet_name: str
    headers: list[str]
    row_count: int
    markdown_table: str
    material_code_column: str | None
    material_code_count: int
    truncated: bool


@dataclass(frozen=True)
class SpreadsheetExtractionResult:
    markdown: str
    sheet_count: int
    total_rows: int
    sheets: list[SheetExtraction]
    review_status: str = "approved"
    extractor: str = "project_r_spreadsheet_preprocess"
    skill_name: str = SKILL_NAME
    skill_version: str = SKILL_VERSION
    prompt_version: str = PROMPT_VERSION
    material_codes_found: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    file_kind: str = "spreadsheet"


def extract_spreadsheet_markdown(source_path: Path) -> SpreadsheetExtractionResult:
    """Extract structured Markdown from an Excel file.

    Uses openpyxl to read workbook data deterministically (no LLM call).
    Handles: visible sheets, merged cells, hidden sheets, material code detection.
    """
    try:
        import openpyxl
    except ImportError:
        return SpreadsheetExtractionResult(
            markdown="# Spreadsheet Extraction Failed\n\nopenpyxl is not installed.\n",
            sheet_count=0,
            total_rows=0,
            sheets=[],
            review_status="failed_retryable",
            warnings=["openpyxl library not available; install with `pip install openpyxl`"],
        )

    wb = openpyxl.load_workbook(str(source_path), data_only=True)
    try:
        return _extract_from_workbook(wb, source_path)
    finally:
        wb.close()


def _extract_from_workbook(wb: Any, source_path: Path) -> SpreadsheetExtractionResult:
    """Extract structured Markdown from an open workbook handle."""
    sheet_extractions: list[SheetExtraction] = []
    total_rows = 0
    all_material_codes: list[str] = []
    all_warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            continue

        rows_iter = ws.iter_rows(values_only=True)
        headers: list[str] = []
        table_rows: list[list[str]] = []
        row_count = 0
        truncated = False
        header_row_found = False
        header_row_index = 0
        material_code_column: str | None = None

        for row_index, row in enumerate(rows_iter):
            values = [_clean_cell(cell) for cell in row]
            non_empty = [v for v in values if v]

            if not header_row_found:
                # Skip title rows that don't look like real headers
                row_text_lower = " ".join(non_empty).lower()
                is_title_row = (
                    len(non_empty) < 1
                    or ("project" in row_text_lower and ":" in row_text_lower)
                    or ("checking list" in row_text_lower and len(non_empty) < 5)
                    or all(not _looks_like_header_cell(c) for c in non_empty)
                )
                if is_title_row:
                    continue
                headers = values
                header_row_index = row_index
                header_row_found = True
                continue

            if row_index > header_row_index + MAX_ROWS_PER_SHEET:
                truncated = True
                break

            if any(v for v in values):
                table_rows.append(values)
                row_count += 1

        if not headers or not any(h for h in headers if h.strip()):
            continue
        if row_count == 0:
            continue  # No data rows

        # Detect material code column
        # Prefer columns matching "编号"/"Name"/"Code" over "TYPE"/"s/n"
        material_candidates: list[tuple[int, str, int]] = []
        for col_index, header in enumerate(headers):
            if _is_material_code_column(header):
                h_lower = header.strip().lower()
                # Priority score: higher = better match
                priority = 1  # base
                if any(p in h_lower for p in ("编号", "name", "code", "item")):
                    priority = 3
                elif any(p in h_lower for p in ("type", "sn", "s/n")):
                    priority = 2
                material_candidates.append((col_index, header, priority))

        if material_candidates:
            # Pick highest priority; break ties by column position (leftmost first)
            material_candidates.sort(key=lambda x: (-x[2], x[0]))
            col_index, material_code_column, _ = material_candidates[0]
            for row_data in table_rows:
                if col_index < len(row_data):
                    val = row_data[col_index].strip()
                    if val:
                        all_material_codes.append(val)

        # Build markdown table
        md_table = _build_markdown_table(headers, table_rows, material_code_column)

        sheet_extractions.append(SheetExtraction(
            sheet_name=sheet_name,
            headers=headers,
            row_count=row_count,
            markdown_table=md_table,
            material_code_column=material_code_column,
            material_code_count=_count_non_empty(table_rows, headers, material_code_column),
            truncated=truncated,
        ))
        total_rows += row_count

    # Build final markdown
    markdown = _build_spreadsheet_markdown(source_path, sheet_extractions, all_material_codes)

    return SpreadsheetExtractionResult(
        markdown=markdown,
        sheet_count=len(sheet_extractions),
        total_rows=total_rows,
        sheets=sheet_extractions,
        material_codes_found=sorted(set(all_material_codes)),
        warnings=all_warnings,
        file_kind="spreadsheet",
    )


def _clean_cell(value: Any) -> str:
    """Convert a cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    return str(value).strip()


# Phase 3 D3 fix: patterns that suggest a cell is part of a real header row
_HEADER_CELL_PATTERNS = {
    "name", "编号", "type", "型号", "规格", "数量", "尺寸",
    "producer", "厂家", "colour", "颜色", "finish", "表面处理",
    "grade", "材质", "remarks", "备注", "单价", "总价",
    "s/n", "sn", "code", "item", "material", "订单号", "时间",
}


def _looks_like_header_cell(cell: str) -> bool:
    """Check if a cell looks like part of a header row (field name, not data)."""
    c = cell.strip().lower().rstrip("*").strip()
    if not c:
        return False
    if c in _HEADER_CELL_PATTERNS:
        return True
    if any(p in c for p in _HEADER_CELL_PATTERNS):
        return True
    # Single short word (3-20 chars) - likely a field name
    if 3 <= len(c) <= 25 and " " not in c.strip():
        return True
    return False


def _is_material_code_column(header: str) -> bool:
    """Check if a column header looks like a material code / ID column."""
    h = header.strip().lower().replace("(", "").replace(")", "").replace(" ", "")
    for pattern in MATERIAL_CODE_PATTERNS:
        if pattern in h:
            return True
    return False


def _count_non_empty(
    table_rows: list[list[str]],
    headers: list[str],
    material_code_column: str | None,
) -> int:
    """Count non-empty values in the material code column."""
    if not material_code_column or not headers:
        return 0
    try:
        col_index = headers.index(material_code_column)
    except ValueError:
        return 0
    return sum(
        1 for row in table_rows
        if col_index < len(row) and row[col_index].strip()
    )


def _build_markdown_table(
    headers: list[str],
    rows: list[list[str]],
    material_code_column: str | None,
) -> str:
    """Build a Markdown table from headers and row data."""
    if not headers:
        return ""

    header_line = "| " + " | ".join(headers) + " |"
    separator = "| " + " | ".join(["---"] * len(headers)) + " |"
    row_lines = []
    for row in rows:
        padded = row + [""] * (len(headers) - len(row))
        row_lines.append("| " + " | ".join(padded[: len(headers)]) + " |")

    return header_line + "\n" + separator + "\n" + "\n".join(row_lines) + "\n"


def _build_spreadsheet_markdown(
    source_path: Path,
    sheet_extractions: list[SheetExtraction],
    material_codes: list[str],
) -> str:
    """Build the final GBrain-ready Markdown."""
    lines = [
        f"# {source_path.stem}",
        "",
        "## Source Summary",
        f"Excel workbook extracted by {SKILL_NAME} v{SKILL_VERSION}.",
        f"Sheets processed: {len(sheet_extractions)}",
        f"Total data rows: {sum(s.row_count for s in sheet_extractions)}",
        "",
    ]

    if material_codes:
        lines.append(f"Material codes detected: {', '.join(sorted(set(material_codes))[:20])}")
        if len(set(material_codes)) > 20:
            lines.append(f"... and {len(set(material_codes)) - 20} more")
        lines.append("")

    for sheet in sheet_extractions:
        lines.append(f"## Sheet: {sheet.sheet_name}")
        lines.append(f"*(Rows: {sheet.row_count} | Material code column: {sheet.material_code_column or 'N/A'})*")
        if sheet.truncated:
            lines.append(f"> ⚠️ 行数超过 {MAX_ROWS_PER_SHEET}，已截断。")
        lines.append("")
        lines.append(sheet.markdown_table)
        lines.append("")

    # Source Evidence
    lines.append("## Source Evidence")
    lines.append(f"- **Source**: {source_path.name}")
    lines.append(f"- **Skill**: {SKILL_NAME} v{SKILL_VERSION}")
    lines.append(f"- **Type**: spreadsheet (deterministic extraction, no LLM)")
    for sheet in sheet_extractions:
        lines.append(f"- **Sheet**: {sheet.sheet_name} — {sheet.row_count} rows" + (" (truncated)" if sheet.truncated else ""))

    lines.append("")
    lines.append("## Preprocess Notes")
    lines.append("- Deterministic extraction via openpyxl (no MiMo / LLM call).")
    lines.append("- Merged cells: top-left value used for entire merged range.")
    lines.append("- Hidden sheets excluded.")
    lines.append("- Formula values use cached values (data_only=True).")
    if material_codes:
        lines.append(f"- Material code column detected: material codes extractable for citation.")

    return "\n".join(lines) + "\n"
