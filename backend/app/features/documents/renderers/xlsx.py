from __future__ import annotations

from pathlib import Path
import re

from app.features.documents.content_parser import DocumentBlock, ParsedDocument, parse_document


def render_xlsx(title: str, content: str, output_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    parsed = parse_document(title, content)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    tables = parsed.tables
    if tables:
        workbook.remove(default_sheet)
        for index, table in enumerate(tables, start=1):
            worksheet = workbook.create_sheet(_sheet_name(parsed, table, index))
            _write_table(worksheet, table.rows)
            _format_sheet(worksheet)
    else:
        default_sheet.title = "内容"
        default_sheet.cell(row=1, column=1, value=parsed.title)
        default_sheet.cell(row=1, column=1).font = Font(bold=True)
        default_sheet.cell(row=2, column=1, value="序号")
        default_sheet.cell(row=2, column=2, value="内容")
        default_sheet.cell(row=2, column=1).font = Font(bold=True)
        default_sheet.cell(row=2, column=2).font = Font(bold=True)
        for row_index, line in enumerate(parsed.plain_lines(), start=3):
            default_sheet.cell(row=row_index, column=1, value=row_index - 2)
            default_sheet.cell(row=row_index, column=2, value=line)
        _format_sheet(default_sheet)
    workbook.save(output_path)
    return output_path


def _sheet_name(parsed: ParsedDocument, table: DocumentBlock, index: int) -> str:
    base = parsed.title if index == 1 else f"表{index}"
    safe = re.sub(r"[\[\]:*?/\\]", "_", base).strip() or f"Sheet{index}"
    return safe[:31]


def _write_table(worksheet, rows: list[list[str]]) -> None:
    from openpyxl.styles import Font, PatternFill

    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=_coerce_cell_value(value))
            if row_index == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E8EEF8")


def _format_sheet(worksheet) -> None:
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"
    if worksheet.max_row > 1 and worksheet.max_column > 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 48)


def _coerce_cell_value(value: str) -> str | int | float:
    stripped = value.strip().replace(",", "")
    if re.fullmatch(r"-?\d+", stripped):
        return int(stripped)
    if re.fullmatch(r"-?\d+\.\d+", stripped):
        return float(stripped)
    return value
